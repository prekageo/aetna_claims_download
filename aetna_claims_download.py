#!/usr/bin/env python3

import json
import openpyxl
import os
import shutil
import sqlite3
import sys
import urllib.request

FIREFOX_PROFILE_PATH = ''
AETNA_MEMBER_ID = ''

class WebBrowser:
    def __init__(self, referer, domain):
        self.referer = referer
        self.cookies = self.get_firefox_cookies(domain)

    def get_firefox_cookies(self, domain):
        conn = sqlite3.connect(f'{FIREFOX_PROFILE_PATH}/cookies.sqlite')
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute('select * from moz_cookies where host = ?', (domain,))
        cookies = {}
        for row in cur.fetchall():
            cookies[row['name']] = row['value']
        return cookies

    def prep_request(self, url):
        cookies = ['%s=%s' % (k,v) for (k,v) in self.cookies.items()]
        cookies = '; '.join(cookies)

        r = urllib.request.Request(url)
        r.add_header('Accept', '*/*')
        r.add_header('Accept-Language', 'en-US,en;q=0.5')
        r.add_header('Connection', 'keep-alive')
        r.add_header('Referer', self.referer)
        r.add_header('User-Agent', 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:63.0) Gecko/20100101 Firefox/63.0')
        r.add_header('Cookie', cookies)
        return r

    def get(self, url, headers = None):
        r = self.prep_request(url)
        if headers:
            for k,v in headers.items():
                r.add_header(k,v)
        return urllib.request.urlopen(r)

def get_tokens():
    line = input('type the tokens: ')
    print('got the tokens', len(line))
    data = dict(x.split('=') for x in line[1:].split('&'))
    return data['access_token'], data['id_token']

def main():
    if os.path.exists('aetna.bak.xlsx'):
        print('File \'aetna.bak.xlsx\' already exists. Not overwriting.', file=sys.stderr)
        sys.exit(1)

    access_token, id_token = get_tokens()
    browser = WebBrowser('https://health.aetna.com/', '.aetna.com')

    url = f'https://apih1.aetna.com/healthcare/prod/v3/aetnahealth_memberships/{AETNA_MEMBER_ID}/consolidatedclaims?claimType=MED,DEN,PHAR&claimStatus=CMPL|RVSD|DENY|Paid|Final|Denied'
    headers = {
        'authorization': f'Bearer {access_token}',
        'id_token': id_token,
    }
    data = browser.get(url, headers = headers).read()
    data = json.loads(data)

    existing_claims = {}
    filename = 'aetna.xlsx'

    if os.path.exists(filename):
        shutil.copy(filename, 'aetna.bak.xlsx')
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        for i, row in enumerate(ws.rows):
            existing_claims[row[0].value] = i + 1
    else:
        wb = openpyxl.Workbook()
        row = [
            'id',
            'type',
            'externalClaimId',
            'status',
            'isPaid',
            'totalSubscriber',
            'dateOfServiceBegin',
            'patient',
            'providerId',
            'legalOwnerName',
            'patientRelationshipToSubscriber',
            'isPayable',
            'payableReason',
            'claimCurrentBalance',
            'adjudicationSource',
            'externalClaimStatus',
            'hasClaimDetails',
            'totalPayable',
            'totalBilled',
        ]
        ws = wb.active
        ws.append(row)
    fill_color = openpyxl.styles.fills.PatternFill(patternType='solid', fill_type='solid', fgColor=openpyxl.styles.Color('ffee88'))

    for claim in sorted(data['readConsolidatedClaimsResponse']['consolidatedclaims'], key = lambda c: c['dateOfServiceBegin']):
        row = [
            claim['id'],
            claim['type'],
            claim['externalClaimId'],
            claim['status'],
            claim['isPaid'],
            claim['totalSubscriber'],
            claim['dateOfServiceBegin'],
            claim['patient']['name']['first'],
            claim['medicalDental']['providerId'],
            claim['medicalDental']['legalOwnerName'],
            claim['medicalDental']['patientRelationshipToSubscriber'],
            claim['isPayable'],
            ','.join(claim['payableReason']),
            claim['claimCurrentBalance'],
            claim['adjudicationSource'],
            claim['externalClaimStatus'],
            claim['hasClaimDetails'],
            claim['totalPayable'],
            claim['totalBilled'],
        ]
        existing_row = existing_claims.get(claim['id'], 0)
        append = False
        if existing_row > 0:
            for i in range(len(row)):
                existing = ws.cell(existing_row, i + 1).value
                if existing == None:
                    existing = ''
                if row[i] != existing:
                    print('claim changed', claim['id'], (i, row[i], existing))
                    append = True
        else:
            append = True

        if append:
            ws.append(row)
            ws.row_dimensions[ws._current_row].fill = fill_color
            for i in range(1,len(row)+1):
                cell = ws.cell(ws._current_row, i)
                cell.fill = fill_color
                if ws._current_row > 1:
                    cell.number_format = ws.cell(ws._current_row - 1, i).number_format

    wb.save(filename)

if __name__ == '__main__':
    main()
